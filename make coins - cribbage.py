import concurrent.futures
import itertools
import os
import psutil
import re
import subprocess
import threading
import time
import json
import requests
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from appium import webdriver
from selenium.webdriver.common.actions.pointer_input import PointerInput
from selenium.webdriver.common.actions.action_builder import ActionBuilder
from selenium.webdriver.common.actions import interaction
from selenium.webdriver.common.action_chains import ActionChains
from appium.webdriver.common.appiumby import AppiumBy
from appium.options.common.base import AppiumOptions
from appium import webdriver
from appium.options.android import UiAutomator2Options
from time import sleep
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.remote.webelement import WebElement
from datetime import datetime, timezone
import queue
import logging
from PIL import Image
import io
import schedule

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

ldconsole_path = 'ldconsole.exe'
config_path = 'config.json'

config: dict = {}
config_lock = threading.Lock()


def init_config():
    global config
    with config_lock:
        with open(config_path) as f:
            config = json.load(f)


def save_config():
    global config
    with config_lock:
        with open(config_path, 'w') as f:
            json.dump(config, f, indent=4)


init_config()


def convert_to_float(s: str):
    s = s.replace(",", "").lower()
    if 'k' in s:
        return float(s.replace('k', '')) * 1e3
    if 'm' in s:
        return float(s.replace('m', '')) * 1e6
    return float(s)


def find_process_by_port(port):
    for conn in psutil.net_connections(kind='inet'):
        if conn.laddr.port == port:
            try:
                return psutil.Process(conn.pid).pid
            except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                pass
    return None


coin_data_queue = queue.Queue()
coin_file_lock = threading.Lock()
COIN_FILENAME = "coin_balance.xlsx"
COIN_FILENAME_TEMP = "coin_balance - copy.xlsx"
coin_consumer_thread: threading.Thread = None
reset_log_file_thread: threading.Thread = None


def backup_coin_excel():
    import shutil
    with coin_file_lock:
        shutil.copy(COIN_FILENAME, COIN_FILENAME_TEMP)

schedule.every(15).minutes.do(backup_coin_excel)

def coin_balance_consumer():
    while True:
        instance_name, package_name, balance = coin_data_queue.get()
        if instance_name is None:
            break
        save_coin_balance(instance_name, package_name, balance, COIN_FILENAME)
        coin_data_queue.task_done()


def save_to_queue(instance_name: str, package_name: str, balance: str):
    coin_data_queue.put((instance_name, package_name, balance))


def get_fill_color(balance: float) -> PatternFill:
    if balance >= 10000:
        return PatternFill(start_color="0029FF52", end_color="0029FF52", fill_type="solid")
    elif balance <= 0:
        return PatternFill(start_color="00FF6629", end_color="00FF6629", fill_type="solid")
    else:
        start_r, start_g, start_b = 0xFF, 0x66, 0x29
        end_r, end_g, end_b = 0x29, 0xFF, 0x52
        try:
            ratio = balance / 10000
            b = int(ratio)
        except:
            ratio = 0
        r = int(start_r + (end_r - start_r) * ratio)
        g = int(start_g + (end_g - start_g) * ratio)
        b = int(start_b + (end_b - start_b) * ratio)
        color_hex = f"00{r:02X}{g:02X}{b:02X}"
        return PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")


def save_coin_balance(instance_name: str, package_name: str, balance: str, filename: str = "coin_balance.xlsx"):
    package_column = f"{instance_name} - Package Name"
    balance_column = f"{instance_name} - Balance"
    updated_at_column = f"{instance_name} - Updated At"
    summary_column = "Summary"
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    with coin_file_lock:
        if os.path.exists(filename):
            df = pd.read_excel(filename)
        else:
            df = pd.DataFrame()
        if package_column not in df.columns:
            df[package_column] = pd.Series(dtype=object)
            df[balance_column] = pd.Series(dtype=object)
            df[updated_at_column] = pd.Series(dtype=object)
        if package_name in df[package_column].values:
            row_index = df[df[package_column] == package_name].index[0]
            df.at[row_index, balance_column] = convert_to_float(balance)
            df.at[row_index, updated_at_column] = current_time
        else:
            num_entries_in_instance = df[package_column].count()
            df.at[num_entries_in_instance, package_column] = package_name
            df.at[num_entries_in_instance,
                  balance_column] = convert_to_float(balance)
            df.at[num_entries_in_instance, updated_at_column] = current_time
        df.to_excel(filename, index=False)

        wb = load_workbook(filename)
        ws = wb.active

        for col_name in df.columns:
            if "Balance".lower() in col_name.lower():
                col_idx = df.columns.get_loc(col_name) + 1
                col_letter = get_column_letter(col_idx)
                for row in range(2, len(df) + 2):
                    balance_value = df.at[row - 2, col_name]
                    fill = get_fill_color(balance_value)
                    ws[f"{col_letter}{row}"].fill = fill

        # Check if "Summary" column already exists
        summary_col_idx = None
        for col in ws.iter_cols(min_row=1, max_row=1):
            if col[0].value == summary_column:
                summary_col_idx = col[0].column
                break

        # If "Summary" column doesn't exist, add it as a new column
        if summary_col_idx is None:
            summary_col_idx = len(ws[1]) + 1
            ws.cell(row=1, column=summary_col_idx, value=summary_column)

        summary_col_letter = get_column_letter(summary_col_idx)

        # Clear existing summary data if present
        for row in ws.iter_rows(min_row=2, min_col=summary_col_idx, max_col=summary_col_idx):
            row[0].value = None

        # 1. Total balances of all instances
        total_balance = df[[
            col for col in df.columns if "Balance" in col]].sum().sum()
        ws[f"{summary_col_letter}2"] = f"All: {total_balance}"

        # 2. Total balances for each instance
        row_offset = 3
        for col_name in df.columns:
            if "Balance" in col_name:
                instance_total = df[col_name].sum()
                ws[f"{summary_col_letter}{row_offset}"] = f"{col_name.split(' - ')[0]} - {instance_total}"
                row_offset += 1

        sum_counts = 0
        # 3. Count balances by range
        for lower in range(0, 10001, 500):
            upper = lower + 499 if lower < 10000 else float('inf')
            count = sum(
                (df[col] >= lower) & (df[col] <= upper)
                for col in df.columns if "Balance" in col
            ).sum()
            sum_counts += count
            range_text = f"{lower} and {upper}: {count}" if upper != float(
                'inf') else f"Above 10000: {count}"
            ws[f"{summary_col_letter}{row_offset}"] = range_text
            row_offset += 1
        
        ws[f"{summary_col_letter}{row_offset}"] = f"total accounts: {sum_counts}"
        row_offset += 1

        wb.save(filename)


def start_consumer_thread():
    global coin_consumer_thread
    coin_consumer_thread = threading.Thread(
        target=coin_balance_consumer, daemon=True)
    coin_consumer_thread.start()


def stop_consumer_thread():
    global coin_consumer_thread
    coin_data_queue.put((None, None, None))
    coin_consumer_thread.join()


def list_ldplayer_instances():
    command = f'"{ldconsole_path}" list2'
    result = subprocess.run(command, stdout=subprocess.PIPE, shell=True)
    instances = result.stdout.decode().splitlines()
    appium_port = 4723
    system_port = 8200
    adb_port = 5556

    instance_names = []
    for line in instances:
        line = line.split(',')
        if line[0] not in config['instances_index'].split(','):
            appium_port += 1
            system_port += 1
            adb_port += 1
            continue
        instance_names.append({
            "index": line[0],
            "name": line[1],
            "status": line[5] != "-1",
            "appium_port": appium_port,
            "system_port": system_port,
            "adb_port": adb_port,
        })
        appium_port += 1
        system_port += 1
        adb_port += 1
    return instance_names


def launch_ldplayer_instance_by_name(instance_name, adb_port):
    subprocess.Popen(
        f'"{ldconsole_path}" launch --name {instance_name} --adb-port {adb_port}')


def launch_ldplayer_instance_by_index(instance_index, adb_port):
    h = "--headless --mute" if config['headless'] != 0 else ''
    subprocess.Popen(
        f'"{ldconsole_path}" launch --index {instance_index} --adb-port {adb_port} {h}')


def quit_ldplayer_instance_by_name(instance_name):
    subprocess.Popen(f'"{ldconsole_path}" quit --name {instance_name}')


def quit_ldplayer_instance_by_index(instance_index):
    subprocess.Popen(f'"{ldconsole_path}" quit --index {instance_index}')


def get_offline_devices():
    result = subprocess.run(['adb', 'devices'], capture_output=True, text=True)
    lines = result.stdout.strip().split('\n')
    offline_devices = []
    for line in lines[1:]:
        if 'offline' in line:
            device_id = line.split()[0]
            offline_devices.append(device_id)
    return offline_devices


def terminate_emulator(emulator):
    subprocess.run(["adb", "-s", emulator, "emu", "kill"],
                   capture_output=True, text=True)


def terminate_all_offline_emulators():
    emulators = get_offline_devices()
    if emulators:
        for emulator in emulators:
            print(f"Terminating {emulator}...")
            terminate_emulator(emulator)


def list_adb_devices(adb_port):
    result = subprocess.run(f'adb devices', stdout=subprocess.PIPE, shell=True)
    devices = result.stdout.decode().splitlines()
    device_ids = [line.split('\t')[0]
                  for line in devices if '\tdevice' in line]
    return device_ids


def wait_for_new_LDPlayer_instance_to_appear_as_a_device(adb_port, timeout=60):
    start_time = time.time()
    logging.info(
        "Waiting for the new LDPlayer instance to appear as a device...")
    initial_devices = set(list_adb_devices(adb_port))
    while time.time() - start_time < timeout:
        current_devices = set(list_adb_devices(adb_port))
        new_devices = current_devices - initial_devices
        if new_devices:
            device_id = list(new_devices)[0]
            logging.info(
                f"LDPlayer instance is running with device ID: {device_id}")
            return device_id
        time.sleep(2)
    raise Exception('time out waiting for LDPlayer instance to launch')


def wait_for_device_ready(device_id, adb_port, timeout=60):
    start_time = time.time()
    while time.time() - start_time < timeout:
        # Check if device is online and fully booted
        boot_completed = subprocess.run(f'adb -s {device_id} shell getprop sys.boot_completed',
                                        stdout=subprocess.PIPE, shell=True)
        if '1' in boot_completed.stdout.decode().strip():
            logging.info(f"Device {device_id} is fully booted.")
            return True
        logging.info(f"Waiting for device {device_id} to boot...")
        time.sleep(5)  # Wait for 5 seconds before checking again
    raise Exception('device could not be ready in time')


def list_installed_plato(device_id, adb_port):
    result = subprocess.run(f'adb -s {device_id} shell pm list packages',
                            stdout=subprocess.PIPE, shell=True)
    packages = result.stdout.decode().splitlines()

    return [x.replace('package:', '') for x in packages if x.startswith('package:com.plato.')]


def start_appium_session(appium_port, system_port, adb_port, device_id, packageName, app_activity):
    desired_caps = {
        'platformName': 'Android',
        'deviceName': device_id,
        'systemPort': system_port,
        'udid': device_id,
        'appPackage': packageName,
        'appActivity': app_activity,
        'automationName': 'UiAutomator2',
        'autoGrantPermissions': True,
        'noReset': True,
        'fullReset': False,
        'newCommandTimeout': 600,
        'adbExecTimeout': 30000,
    }

    driver = webdriver.Remote(
        f'http://localhost:{appium_port}', options=UiAutomator2Options().load_capabilities(desired_caps))
    return driver


def handle_system_ui_not_responding(d: webdriver.Remote):
    try:
        close_app_button = d.find_element(AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().text("Close app")'
        )
        close_app_button.click()
    except Exception:
        pass


def mute_ld_player(d: webdriver.Remote):
    try:
        for _ in range(10):
            d.press_keycode(25)
            time.sleep(0.1)
    except Exception as e:
        pass


def is_game_favorite(d: webdriver.Remote, game_name: str):
    go_to_home_tab(d)
    fs = WebDriverWait(d, 30).until(EC.visibility_of_element_located(
        (By.ID, 'favorites_recycler_view')))
    f = fs.find_elements(By.CLASS_NAME, 'android.view.ViewGroup')[0]
    c_f = f.find_element(By.ID, 'title_text_view').text
    if c_f.lower() == game_name.lower():
        return True
    return False


def is_rank_game_played(d: webdriver.Remote):
    sleep(0.4)
    go_to_shop_tab(d)
    sleep(0.4)
    go_to_home_tab(d)
    sleep(0.4)
    s = time.time()
    while 1:
        try:
            go_to_shop_tab(d)
            sleep(0.4)
            go_to_home_tab(d)
            sleep(0.4)
            ss = time.time()
            while 1:
                try:
                    won_games = WebDriverWait(d, 5).until(EC.visibility_of_element_located(
                        (By.ID, 'quest_progress_bar'))).text
                    break
                except:
                    if time.time() - ss > 30:
                        raise Exception("can't find progress bar")
                    ff = False
                    try:
                        assert "/" in d.find_element(By.ID,
                                                     "plato_conversation_time").text
                        ff = True
                    except Exception as e:
                        try:
                            assert "/" in d.find_element(By.ID,
                                                         "game_type_time").text
                            ff = True
                        except Exception as e:
                            pass
                        pass
                    if ff:
                        raise Exception("couldn't find ranked")
                    for _ in range(4):
                        d.press_keycode(20)
            break
        except:
            if time.time() - s > 90:
                raise Exception("can't find progress bar")
            sleep(0.3)
    try:
        won_games = float(won_games)
        return won_games == 1
    except:
        pass
    return False


def is_game_in_favorite(d: webdriver.Remote, game_name: str):
    go_to_shop_tab(d)
    sleep(0.5)
    go_to_home_tab(d)
    sleep(0.5)
    while 1:
        flag = False
        try:
            recycler = WebDriverWait(d, 5).until(EC.visibility_of_element_located(
                (By.ID, 'favorites_recycler_view')))
        except:
            return False
        for title_element in recycler.find_elements(By.ID, 'title_text_view'):
            if game_name.lower() == title_element.text.lower().strip():
                x = title_element.location_in_view['x'] + \
                    title_element.size['width']//2
                y = title_element.location_in_view['y'] - 50
                d.tap([(x, y)])
                flag = True
                break
        if flag:
            return True
        else:
            size = d.get_window_size()
            x1 = int(size['width'] * 0.75)
            y1 = int(size['height'] * 0.25)
            x2 = int(size['width'] * 0.35)
            y2 = int(size['height'] * 0.25)
            d.swipe(x1, y1, x2, y2, 150)


def select_game(d: webdriver.Remote, game_name: str):
    if is_game_in_favorite(d, game_name):
        return

    game_tab = WebDriverWait(d, 10).until(EC.visibility_of_element_located(
        (By.ID, 'plato_image_games')))
    game_tab.click()
    game_button = WebDriverWait(d, 10).until(EC.visibility_of_element_located(
        (By.ID, 'game_list_item_container')))
    found_games = []
    while 1:
        games_button = d.find_elements(
            By.ID, "game_list_item_container")
        found_new = False
        for game in games_button:
            try:
                game_title = game.find_element(
                    By.ID, 'game_list_item_title').text
                if game_title not in found_games:
                    if game_name.lower() == game_title.lower():
                        game.click()
                        toggle_favorite(d)
                        return True
                    found_games.append(game_title)
                    found_new = True
            except Exception as e:
                pass
        else:
            found_new = True
        if not found_new:
            break
        for _ in range(4):
            d.press_keycode(20)
    return False


def is_game_closed(d: webdriver.Remote):
    screenshot = d.get_screenshot_as_png()
    image = Image.open(io.BytesIO(screenshot))
    width, height = image.size
    x = int(width * 0.5)
    y = int(height * 0.5)
    pixel_color = image.getpixel((x, y))
    white = (255, 255, 255)

    def color_distance(c1, c2):
        return sum((a - b) ** 2 for a, b in zip(c1, c2)) ** 0.5
    res = 50 > color_distance(pixel_color, white)
    return res


def cribbage_is_my_turn(d: webdriver.Remote):
    "when it's yellow, it's my turn: returns True"
    screenshot = d.get_screenshot_as_png()
    image = Image.open(io.BytesIO(screenshot))
    width, height = image.size
    x = int(width * 0.85)
    y = int(height * 0.485)
    pixel_color = image.getpixel((x, y))
    yellow = (255, 255, 0)
    blue = (0, 0, 255)

    def color_distance(c1, c2):
        return sum((a - b) ** 2 for a, b in zip(c1, c2)) ** 0.5
    distance_to_yellow = color_distance(pixel_color, yellow)
    distance_to_blue = color_distance(pixel_color, blue)
    if distance_to_yellow < distance_to_blue:
        # logging.info("it's my turn.")
        return True
    else:
        # logging.info("it's not my turn.")
        return False


def play_latest_rank_season(d: webdriver.Remote):
    WebDriverWait(d, 10).until(
        EC.visibility_of_element_located((By.ID, 'enterable_item_title')))
    s = time.time()
    while len(d.find_elements(By.ID, "enterable_item_title")) < 2:
        sleep(0.5)
        if time.time() - s > 30:
            raise Exception(
                "couldn't find ranked season matchmaking (maybe net problem)")
    close_previous_games(d)
    found_matchmaking_buttons: list[tuple[WebElement, str]] = []
    s = time.time()
    flag = False
    while not flag:
        for matchmaking_title in d.find_elements(By.ID, "enterable_item_title"):
            try:
                txt = matchmaking_title.text
                if 'rank' in txt.lower():
                    flag = True
            except Exception as e:
                pass
        if time.time() - s > 30:
            raise Exception(
                "couldn't find ranked season matchmaking (maybe net problem)")
    s = time.time()
    while 1:
        while 1:
            found_new = False
            for matchmaking_title in d.find_elements(By.ID, "enterable_item_title"):
                try:
                    txt = matchmaking_title.text
                    if 'rank' in txt.lower() and txt not in [x[1] for x in found_matchmaking_buttons]:
                        found_matchmaking_buttons.append([matchmaking_title, txt, int(
                            re.findall(r'\d+', txt.lower().split('season')[1])[0])])
                        found_new = True
                except Exception as e:
                    pass
            if not found_new:
                break
            for _ in range(4):
                d.press_keycode(20)
        if len(found_matchmaking_buttons) > 0:
            max(found_matchmaking_buttons, key=lambda x: x[2])[0].click()
            break
        else:
            if time.time() - s > 30:
                raise Exception(
                    "couldn't find ranked season matchmaking (maybe net problem)")
    WebDriverWait(d, 10).until(
        EC.element_to_be_clickable((By.ID, 'join_button'))).click()
    sleep(3)
    s = time.time()
    while 1:
        try:
            el = d.find_element(By.ID, 'enterable_item_message')
            txt = el.text
            assert 'match made' in txt.lower()
            el.click()
            break
        except:
            if time.time() - s > 10:
                raise Exception(
                    "couldn't find ranked season matchmaking (maybe net problem)")
        sleep(0.5)
    sleep(3)
    WebDriverWait(d, 3*60).until(EC.invisibility_of_element_located((By.ID,
                                                                     'plato_container_game_spinner')))
    sleep(1)


def close_previous_games(d: webdriver.Remote):
    try:
        WebDriverWait(d, 5).until(EC.visibility_of_element_located(
            (By.XPATH, '//android.widget.TextView[@text="PLAY"] | //android.widget.TextView[@text="Searching…"]')))
    except:
        return
    retry = 5
    while 1:
        retry -= 1
        try:
            d.find_element(
                By.XPATH, '//android.widget.TextView[@text="Searching…"]').click()
            WebDriverWait(d, 2).until(EC.visibility_of_element_located(
                (By.XPATH, '//android.widget.Button[@text="REMOVE"]'))).click()
            sleep(1)
        except:
            try:
                d.find_element(
                    By.XPATH, '//android.widget.TextView[@text="PLAY"]').click()
                WebDriverWait(d, 3*60).until(EC.invisibility_of_element_located((By.ID,
                                                                                'plato_container_game_spinner')))
                d.back()
                if retry <= 0:
                    raise Exception(
                        "couldn't close previous games")
                sleep(1)
            except:
                break
        sleep(0.5)


def tap_using_percent(d: webdriver.Remote, x_percent: float, y_percent: float):
    window_size = d.get_window_size()
    screen_width = window_size['width']
    screen_height = window_size['height']
    x = int(screen_width * x_percent)
    y = int(screen_height * y_percent)
    d.tap([(x, y)], 10)


def resign_from_game(d:  webdriver.Remote):
    WebDriverWait(d, 10).until(EC.visibility_of_element_located(
        (By.ID, 'plato_button_hamburger'))).click()
    WebDriverWait(d, 10).until(EC.visibility_of_element_located(
        (AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().text("Resign")'))).click()
    sleep(1)
    size = d.get_window_size()
    if config['win_fake_game'].lower() in ['match monsters']:
        tap_using_percent(d, 0.25, 0.40)
        tap_using_percent(d, 0.25, 0.45)
        tap_using_percent(d, 0.25, 0.485)
        tap_using_percent(d, 0.25, 0.52)
    elif config['win_fake_game'].lower() in ['ludo']:
        tap_using_percent(d, 0.25, 0.53)
        tap_using_percent(d, 0.25, 0.57)
        tap_using_percent(d, 0.25, 0.6)
        tap_using_percent(d, 0.25, 0.63)
    elif config['win_fake_game'].lower() in ['dots & boxes']:
        tap_using_percent(d, 0.75, 0.6)
    else:
        if size['width'] > size['height']:
            tap_using_percent(d, 0.6, 0.79)
        else:
            tap_using_percent(d, 0.65, 0.53)
            tap_using_percent(d, 0.75, 0.60)
            tap_using_percent(d, 0.75, 0.65)
            tap_using_percent(d, 0.75, 0.685)
            tap_using_percent(d, 0.75, 0.72)
    sleep(0.7)
    d.back()


def toggle_favorite(d:  webdriver.Remote):
    WebDriverWait(d, 10).until(EC.visibility_of_element_located(
        (AppiumBy.ACCESSIBILITY_ID, 'Added to Favorites'))).click()


def get_coins(d: webdriver.Remote):
    go_to_shop_tab(d)
    balance = WebDriverWait(d, 10).until(EC.visibility_of_element_located(
        (By.ID, 'wallet_view_balance'))).text
    return balance


def go_to_shop_tab(d: webdriver.Remote):
    WebDriverWait(d, 10).until(EC.visibility_of_element_located(
        (By.ID, 'plato_tab_shop'))).click()


def go_to_home_tab(d: webdriver.Remote):
    WebDriverWait(d, 10).until(EC.visibility_of_element_located(
        (By.ID, 'plato_tab_home'))).click()


def run_appium_server(appium_server_port):
    # run_appium_server_non_headless(appium_server_port)
    run_appium_server_headless(appium_server_port)


def run_appium_server_non_headless(appium_server_port):
    while 1:
        command = f'start cmd /c "appium --relaxed-security -p {appium_server_port} && exit"'
        process = subprocess.Popen(command, shell=True)
        server_url = f'http://localhost:{appium_server_port}/status'
        start_time = time.time()
        timeout = 30
        while time.time() - start_time < timeout:
            try:
                response = requests.get(server_url)
                if response.status_code == 200:
                    logging.info("Appium server is running!")
                    return process
            except Exception:
                pass
            time.sleep(1)
    raise Exception("Appium server failed to start within the timeout period.")


def run_appium_server_headless(appium_server_port):
    while 1:
        command = f'appium --relaxed-security -p {appium_server_port}'
        process = subprocess.Popen(
            command, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, shell=True)
        server_url = f'http://localhost:{appium_server_port}/status'
        start_time = time.time()
        timeout = 30
        while time.time() - start_time < timeout:
            try:
                response = requests.get(server_url)
                if response.status_code == 200:
                    logging.info("Appium server is running!")
                    return process
            except Exception:
                pass
            time.sleep(1)
    raise Exception("Appium server failed to start within the timeout period.")


def stop_appium_server(appium_server_port):
    pid = find_process_by_port(appium_server_port)
    if pid:
        try:
            process = psutil.Process(pid)
            process.terminate()
            process.wait(timeout=5)
            logging.info(
                f"Appium server on port {appium_server_port} has been stopped.")
        except Exception as e:
            logging.error(f"Failed to stop Appium server: {e}")
    else:
        logging.info(f"No process found on port {appium_server_port}")


def click_lets_go(d: webdriver.Remote):
    try:
        WebDriverWait(d, 3).until(EC.visibility_of_element_located(
            (By.ID, 'start_screen_button_label'))).click()
    except:
        pass
    WebDriverWait(d, 20).until(EC.visibility_of_element_located(
        (By.ID, 'plato_tab_home')))


def check_for_backup_button(d: webdriver.Remote):
    try:
        WebDriverWait(d, 3).until(EC.visibility_of_element_located(
            (By.XPATH, '//android.widget.Button[@text="CANCEL"]'))).click()
    except:
        pass


launch_instance_appium_server_lock = threading.Lock()


def launch_instance(instance: dict):
    instance_index = instance["index"]
    instance_name = instance["name"]
    instance_adb_port = instance["adb_port"]
    while 1:
        retry = 4
        with launch_instance_appium_server_lock:
            while retry > 0:
                try:
                    logging.info(
                        f"Launching LDPlayer instance: {instance_name}")
                    launch_ldplayer_instance_by_index(
                        instance_index, instance_adb_port)
                    device_id = wait_for_new_LDPlayer_instance_to_appear_as_a_device(
                        instance_adb_port)
                    wait_for_device_ready(device_id, instance_adb_port)
                    return device_id
                except Exception as e:
                    retry -= 1
                    logging.error(
                        f"failed to launch LDPlayer instance: {instance_name} --------------")
                    try:
                        quit_ldplayer_instance_by_index(instance_index)
                    except:
                        pass
                    terminate_all_offline_emulators()
        if retry > 0:
            break
        else:
            sleep(10)


done_instances = []


def run_instance(instance: dict):
    app_activity = 'com.playchat.ui.activity.MainActivity'
    instance_index = instance["index"]
    instance_name = instance["name"]
    instance_appium_port = instance["appium_port"]
    instance_system_port = instance["system_port"]
    instance_adb_port = instance["adb_port"]
    device_id = launch_instance(instance)
    d: webdriver.Remote = None

    def safe_quit():
        try:
            d.quit()
        except:
            pass
        stop_appium_server(instance_appium_port)
        quit_ldplayer_instance_by_index(instance_index)
        sleep(5)

    logging.info(f"Starting Appium Server for instance: {instance_name}")
    run_appium_server(instance_appium_port)
    installed_platos = list_installed_plato(device_id, instance_adb_port)
    d = start_appium_session(instance_appium_port, instance_system_port,
                                instance_adb_port, device_id, installed_platos[0], app_activity)
    mute_ld_player(d)
    try:
        d.quit()
    except:
        pass
    
    import random
    random.shuffle(installed_platos)

    instance_is_done = True

    for package_name in installed_platos:
        if is_processed_app_logged(instance_index, package_name):
            continue
        retry = 3
        while 1:
            error_type = 1
            try:
                logging.info(
                    f"Starting Appium session on the device on instance {instance_name} for {package_name} app")
                d = start_appium_session(instance_appium_port, instance_system_port,
                                         instance_adb_port, device_id, package_name, app_activity)
                handle_system_ui_not_responding(d)
                error_type = 2
                logging.info(f"launching app {package_name}")
                d.activate_app(package_name)
                click_lets_go(d)
                # check_for_backup_button(d)
                if not is_rank_game_played(d):
                    instance_is_done = False
                    select_game(d, 'Cribbage')
                    play_latest_rank_season(d)
                    sleep(5)
                    if is_game_closed(d):
                        d.back()
                    else:
                        if cribbage_is_my_turn(d):
                            d.back()
                        else:
                            resign_from_game(d)
                    d.back()
                else:
                    logging.info(
                        f"on instance {instance_name} for {package_name} app quest is done")
                    log_processed_app(instance_index, package_name)
                balance = get_coins(d)
                save_to_queue(instance_name, package_name, balance)
                d.terminate_app(package_name)
                d.quit()
                break
            except Exception as e:
                retry -= 1
                if error_type == 1:
                    safe_quit()
                    sleep(2)
                    device_id = launch_instance(instance)
                    run_appium_server(instance_appium_port)
                if error_type == 2:
                    logging.error(
                        f"failed to launch app on instance {instance_name} for {package_name} -------------- {e}")
                    try:
                        d.terminate_app(package_name)
                        d.quit()
                    except:
                        pass
                if retry <= 0:
                    break

                # if retry <= 0:
                #     safe_quit()
                #     return instance_index
                # safe_quit()
                # sleep(2)
                # device_id = launch_instance(instance)
                # run_appium_server(instance_appium_port)
    safe_quit()
    if not instance_is_done:
        return run_instance(instance)
    else:
        done_instances.append(instance_index)
    return instance


def chunk_list(input_list, chunk_size=4):
    return [input_list[i:i + chunk_size] for i in range(0, len(input_list), chunk_size)]


def get_file_done_path_for_instance(instance_index: str):
    return "done/" + str(instance_index)


def check_for_reset_log_file():
    while True:
        schedule.run_pending()
        sleep(20)


def clear_done_instances():
    global done_instances
    done_instances.clear()


def initialize_log():
    all_instances = list_ldplayer_instances()
    os.makedirs("done/", exist_ok=True)
    for instance in all_instances:
        instance_index = instance["index"]
        try:
            with open(get_file_done_path_for_instance(instance_index), "r") as file:
                json.load(file)
        except (FileNotFoundError, json.JSONDecodeError):
            with open(get_file_done_path_for_instance(instance_index), "w") as file:
                json.dump({}, file)
    schedule.every().day.at("03:30").do(clear_done_instances)
    schedule.every().day.at("03:30").do(reset_log_file)
    th = threading.Thread(target=check_for_reset_log_file, daemon=True)
    th.start()


def is_processed_app_logged(instance_index: str, app_name: str):
    with open(get_file_done_path_for_instance(instance_index), "r") as file:
        data = json.load(file)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if today not in data:
        data[today] = {}
    if instance_index not in data[today]:
        data[today][instance_index] = []
    return app_name in data[today][instance_index]


def log_processed_app(instance_index: str, app_name: str):
    with open(get_file_done_path_for_instance(instance_index), "r") as file:
        data = json.load(file)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if today not in data:
        data[today] = {}
    if instance_index not in data[today]:
        data[today][instance_index] = []
    data[today][instance_index].append(app_name)
    with open(get_file_done_path_for_instance(instance_index), "w") as file:
        json.dump(data, file, indent=4)
    return False


def reset_log_file():
    all_instances = list_ldplayer_instances()
    for instance in all_instances:
        instance_index = instance["index"]
        with open(get_file_done_path_for_instance(instance_index), "w") as file:
            json.dump({}, file)


def main():
    done_instances.clear()
    initialize_log()
    all_instances = list_ldplayer_instances()
    max_workers = min(config['total_launched_instances'], len(all_instances))
    start_consumer_thread()
    while True:
        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = [executor.submit(run_instance, i) for i in all_instances]
            done, undone = concurrent.futures.wait(
                futures, return_when=concurrent.futures.ALL_COMPLETED)
            while len(done_instances) == len(all_instances):
                sleep(1)
    # coin_data_queue.join()
    # stop_consumer_thread()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nProgram interrupted! Exiting...")
